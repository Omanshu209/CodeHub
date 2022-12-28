
import openpyxl

file=openpyxl.load_workbook("Excel.xlsx")
ws=file.active

global r,n,mm,pp,cc
r=3
n=3
mm=3
pp=3
cc=3

class student:
	
	def __init__(self):
		self.r_no=r
		self.n_no=n
		self.m_no=mm
		self.p_no=pp
		self.c_no=cc
		
	def details(self,x,y,m=0,p=0,c=0):
		
		self.x=x
		self.m=m
		self.y=y
		self.p=p
		self.c=c
			
		ws[f"A{self.r_no}"].value=self.x
		ws[f"B{self.n_no}"].value=self.y
		
		if m!=0 or p!=0 or c!=0:
		    ws[f"D{self.m_no}"].value=self.m
		    ws[f"E{self.p_no}"].value=self.p
		    ws[f"F{self.c_no}"].value=self.c
		
		else:
			ws[f"D{self.m_no}"].value="Marks Not Available"
			ws.merge_cells(f"D{self.m_no}:F{self.c_no}")
		
		global mm,pp,cc,r,n
		mm+=1
		pp+=1
		cc+=1
		r+=1
		n+=1		
	
ws["A1"].value="Roll No."
ws["B1"].value="Name"
ws["D1"].value="MATHS"
ws["E1"].value="PHY"
ws["F1"].value="CHEM"

student1=student()
student1.details(7,"IDK",57,63,42)

student2=student()
student2.details(2,"LOL",23,17)

student3=student()
student3.details(9,"PLS")

file.save("Excel.xlsx")
